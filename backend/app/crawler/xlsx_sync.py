"""
xlsx 定期同步模块
每 N 条数据或每 M 分钟同步一次到 xlsx 文件
"""
import logging
import os
import sqlite3
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

from openpyxl import Workbook, load_workbook

logger = logging.getLogger("tiktokrx.crawler.xlsx_sync")

DEFAULT_SYNC_THRESHOLD = 100  # 条数阈值
DEFAULT_TIME_THRESHOLD = 300  # 5分钟 = 300秒


class XlsxSync:
    """xlsx 同步器"""

    def __init__(
        self,
        db_path: str,
        output_dir: str,
        sync_count: int = DEFAULT_SYNC_THRESHOLD,
        sync_interval: int = DEFAULT_TIME_THRESHOLD,
    ):
        self.db_path = db_path
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.sync_count = sync_count
        self.sync_interval = sync_interval

        self._lock = threading.Lock()
        self._last_sync_time = time.time()
        self._pending_count = 0
        self._table_last_sync: Dict[str, int] = {}  # table -> last sync row count

    def _get_tables(self) -> List[str]:
        """获取需要同步的表"""
        return ["video_database", "video_database", "bgm_database", "title_database", "comment_database", "cover_database"]

    def _get_table_count(self, table: str) -> int:
        """获取表记录数"""
        try:
            conn = sqlite3.connect(self.db_path)
            count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            conn.close()
            return count
        except Exception:
            return 0

    def _should_sync(self) -> bool:
        """检查是否需要同步"""
        now = time.time()
        if self._pending_count >= self.sync_count:
            return True
        if now - self._last_sync_time >= self.sync_interval:
            return True
        return False

    def _sync_single_table(self, wb: Workbook, table: str) -> int:
        """同步单个表到 xlsx，返回新增记录数"""
        conn = sqlite3.connect(self.db_path)
        try:
            rows = conn.execute(f"SELECT * FROM {table}").fetchall()
            cols = [desc[0] for desc in conn.execute(f"PRAGMA table_info({table})").fetchall()]
            conn.close()

            last_count = self._table_last_sync.get(table, 0)
            new_rows = rows[last_count:]
            if not new_rows:
                return 0

            # 追加到 sheet
            if table in wb.sheetnames:
                ws = wb[table]
            else:
                ws = wb.create_sheet(table)
                ws.append(cols)

            for row in new_rows:
                ws.append(list(row))

            self._table_last_sync[table] = len(rows)
            return len(new_rows)
        except Exception as e:
            logger.warning(f"同步表 {table} 失败: {e}")
            try:
                conn.close()
            except Exception:
                pass
            return 0

    def sync_to_xlsx(self, force: bool = False) -> Optional[str]:
        """
        同步数据库到 xlsx 文件
        返回 xlsx 文件路径，失败返回 None
        """
        if not force and not self._should_sync():
            return None

        with self._lock:
            xlsx_path = self.output_dir / "抖音原始数据.xlsx"
            try:
                if xlsx_path.exists():
                    wb = load_workbook(xlsx_path)
                else:
                    wb = Workbook()
                    if "Sheet" in wb.sheetnames:
                        del wb["Sheet"]

                total_new = 0
                for table in self._get_tables():
                    count = self._sync_single_table(wb, table)
                    if count > 0:
                        logger.info(f"表 {table} 新增 {count} 条")
                        total_new += count

                if total_new > 0:
                    wb.save(xlsx_path)
                    logger.info(f"xlsx 同步完成: {xlsx_path}，共 {total_new} 条新记录")

                self._last_sync_time = time.time()
                self._pending_count = 0
                return str(xlsx_path)
            except Exception as e:
                logger.error(f"xlsx 同步失败: {e}")
                return None

    def on_records_saved(self, count: int = 1) -> Optional[str]:
        """每保存一批记录后调用，检查是否触发同步"""
        self._pending_count += count
        return self.sync_to_xlsx()

    def flush(self) -> Optional[str]:
        """强制同步所有待写入记录"""
        return self.sync_to_xlsx(force=True)


# 全局同步器实例
_global_syncers: Dict[str, XlsxSync] = {}
_sync_lock = threading.Lock()


def get_xlsx_sync(db_path: str, output_dir: str) -> XlsxSync:
    """获取全局同步器实例（单例）"""
    key = f"{db_path}:{output_dir}"
    with _sync_lock:
        if key not in _global_syncers:
            _global_syncers[key] = XlsxSync(db_path, output_dir)
        return _global_syncers[key]


def sync_all_tables_to_xlsx(
    db_path: str,
    output_dir: str,
    tables: Optional[List[str]] = None,
) -> Optional[str]:
    """
    一次性将指定表同步到 xlsx（不进行增量检查，直接全量导出）
    用于手动触发或定时任务
    """
    output_path = Path(output_dir) / "抖音原始数据.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if tables is None:
        tables = ["video_database", "bgm_database", "title_database", "comment_database", "cover_database"]

    try:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        conn = sqlite3.connect(db_path)
        for table in tables:
            try:
                rows = conn.execute(f"SELECT * FROM {table}").fetchall()
                cols = [desc[0] for desc in conn.execute(f"PRAGMA table_info({table})").fetchall()]
                ws = wb.create_sheet(table)
                ws.append(cols)
                for row in rows:
                    ws.append(list(row))
                logger.info(f"导出表 {table}: {len(rows)} 条")
            except Exception as e:
                logger.warning(f"导出表 {table} 失败: {e}")
        conn.close()

        wb.save(output_path)
        logger.info(f"全量导出完成: {output_path}")
        return str(output_path)
    except Exception as e:
        logger.error(f"全量导出失败: {e}")
        return None


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    db = "data/tiktok_baseline.db"
    out = "data/抖音原始数据"
    sync_all_tables_to_xlsx(db, out)
